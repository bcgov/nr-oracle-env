"""
Retrieve, read, and consolidate data classification information.

First retrieves the primary keys / foreign keys from a database.  Usually this
should be run against the "just enough" oracle database.

Reads through ss pulled from the sharepoint site:
https://bcgov.sharepoint.com/teams/03678/Shared%20Documents/Forms/AllItems.aspx?id=%2Fteams%2F03678%2FShared%20Documents%2FGeneral%2FData%20Classifications%2FForests%5FTest%5FData%5FRefresh%2FData%5FClassification%5FSpreadsheets%2FApproved%5Fby%5FSecurity&viewid=d81a9f32%2Dd7fe%2D490a%2Dbf22%2Dcea9e64a79d5&csf=1&web=1&e=LprZSo&ovuser=6fdb5200%2D3d0d%2D4a8a%2Db036%2Dd3685e359adc%2CKevin%2ENetherton%40gov%2Ebc%2Eca&OR=Teams%2DHL&CT=1752704712726&clickparams=eyJBcHBOYW1lIjoiVGVhbXMtRGVza3RvcCIsIkFwcFZlcnNpb24iOiI0OS8yNTA2MTIxNjQyMSIsIkhhc0ZlZGVyYXRlZFVzZXIiOmZhbHNlfQ%3D%3D&CID=4578b2a1%2D40c2%2D9000%2D550d%2D2664c4b6e918&cidOR=SPO&FolderCTID=0x012000661E056E05762A4C9AEB04E04A706D67


"""

import enum
import json
import logging
import logging.config
import pathlib

import openpyxl
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import oracledb

import config

logger = logging.getLogger(__name__)



class DataClasses(enum.Enum):
    """
    Enum class for data classifications.
    """

    PUBLIC = 1
    A = 2
    B = 3
    C = 4
    CONFIDENTIAL = 5


class DataClassifications:
    """
    A utility class to store data classification information.

    The structure is used to store data classification information will be a
    table_name -> column name - > security_classification mapping.
    """

    def __init__(self) -> None:
        """
        Construct DataClassifications object.
        """
        self.struct = {}
        self.public_security_classification_values = [
            "NONE",
            "",
            "TABLE NOT REQUIRED",
        ]
        self.a_values = [
            "A",
            "PROTECTED A",
            "PROTECTED_A",
        ]
        self.b_values = [
            "B",
            "PROTECTED B",
            "PROTECTED_B",
        ]
        self.c_values = ["C", "PROTECTED C", "PROTECTED_C", "PROTECTED B OR C"]

        self.public_security_classification_values = [
            value.upper()
            for value in self.public_security_classification_values
        ]
        self.a_values = [value.upper() for value in self.a_values]
        self.b_values = [value.upper() for value in self.b_values]
        self.c_values = [value.upper() for value in self.c_values]
        # if the sec class is not set assume public
        self.public_security_classification_values.append(None)

    def add_classification(
        self, table_name: str, column_name_str: str,
        security_classification_str: str,
    ) -> None:
        """Add a data classification to the structure."""
        logger.debug("sc is: %s", security_classification_str)
        security_classification = self.get_security_classification(
            security_classification_str,
        )
        if table_name not in self.struct:
            self.struct[table_name] = {}
        if column_name_str not in self.struct[table_name]:
            self.struct[table_name][column_name_str] = security_classification
        else:
            existing_security_classification = self.struct[table_name][
                column_name_str
            ]
            if (
                existing_security_classification.value
                < security_classification.value
            ):
                logger.warning(
                    "Overwriting existing security classification %s for "
                    "table %s, column %s with %s",
                    existing_security_classification.name,
                    table_name,
                    column_name_str,
                    security_classification.name,
                )
                self.struct[table_name][column_name_str] = (
                    security_classification
                )
            else:
                logger.info(
                    "Not overwriting existing security classification %s for"
                    "table %s, column %s",
                    existing_security_classification.name,
                    table_name,
                    column_name_str,
                )

    def get_security_classification(self,
                                    security_classification_str: str,
                                    ) -> DataClasses:
        """Get the security classification from the string."""
        logger.debug("security classification: %s", security_classification_str)
        try:
            if security_classification_str is not None:
                security_classification_str = (
                    security_classification_str.strip().upper()
                )
            if (
                security_classification_str
                in self.public_security_classification_values
            ):
                logger.debug(
                    "Security classification is public or not set, returning"
                    " PUBLIC",
                )
                return DataClasses.PUBLIC
            if security_classification_str in self.a_values:
                logger.debug(
                    "Security classification value %s, is resolving to A",
                    security_classification_str,
                )
                return DataClasses.A

            if security_classification_str in self.b_values:
                logger.debug(
                    "Security classification value %s, is resolving to B",
                    security_classification_str,
                )
                return DataClasses.B
            if security_classification_str in self.c_values:
                logger.debug(
                    "Security classification value %s, is resolving to C",
                    security_classification_str,
                )
                return DataClasses.C
            return DataClasses[security_classification_str.upper()]
        except KeyError as exc:
            logger.exception(
                "Invalid security classification: %s",
                security_classification_str,
            )
            raise KeyError from exc

    def dump_struct(self, dest_doc: pathlib.Path, indent: int = 2) -> None:
        """Dump the structure to a JSON file."""
        logger.info("Dumping data classification structure to %s", dest_doc)

        struct = {}
        for table_name, columns in self.struct.items():
            struct[table_name] = {}
            for column_name, security_classification in columns.items():
                struct[table_name][column_name] = security_classification.name

        with dest_doc.open("w") as f:
            json.dump(struct, f, indent=indent)
        logger.info("Data classification structure dumped successfully.")


class ConsoldateDC:
    """
    Consolidate data classification information from Excel files.
    """

    def __init__(
        self,
        src_dir: pathlib.Path,
        dest_doc: pathlib.Path,
    ) -> None:
        """
        Construct ConsoldateDC object.
        """

        self.src_dir = src_dir
        self.dest_doc = dest_doc

        self.struct = {}
        self.dc = DataClassifications()

        self.ignore_sheets = []

        # dictionary to link the workbook to the data classification sheet
        # that resides in the workbook... there can be other sheets in the wb,
        # but the one we care about is the data classification sheet
        self.wb_dc_sheet = {}

    def consolidate(self) -> None:
        """
        Consolidate data classification information.

        Reads the files in the source directory, and attempts to parse data
        classification information from all the Excel files in the directory,
        and eventually extract the data classification information and writes
        it to the destination document. (json)
        """

        logger.info("Consolidating data classification information...")
        logger.debug("Source directory: %s", self.src_dir)
        suffix = ".xlsx"
        xl_files = list(self.src_dir.glob(f"*{suffix}"))
        logger.debug("xl files: %s", xl_files)

        for xl_file in xl_files:
            self.read_xl_file(xl_file)

        self.dc.dump_struct(
            self.dest_doc,
            indent=2,
        )

    def read_xl_file(self, xl_file: pathlib.Path) -> None:
        """
        Read an Excel file and return its content.
        """

        logger.info("Reading Excel file: %s", xl_file)
        wb = openpyxl.load_workbook(xl_file, read_only=True)
        sheet_names = wb.sheetnames
        logger.debug("Sheet names: %s", sheet_names)
        wb_verified = False
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            logger.debug("Processing sheet: %s", sheet)
            if self.has_data(sheet):
                # Process the worksheet if it has data
                logger.debug("Processing sheet: %s", sheet_name)
                # verifies that it has a re
                if self.verify_sheet(sheet):
                    wb_verified = True
                    if xl_file.name not in self.wb_dc_sheet:
                        self.wb_dc_sheet[xl_file.name] = []
                    self.wb_dc_sheet[xl_file.name].append(sheet_name)

        if not wb_verified:
            logger.error(
                "No valid worksheet found in workbook %s with required columns",
                xl_file,
            )
            raise ValueError
        for sheet_name in self.wb_dc_sheet[xl_file.name]:
            sheet = wb[sheet_name]
            logger.debug("Extracting data from sheet: %s", sheet_name)
            # Extract data from the verified sheet
            self.extract_data(sheet, sheet_name)

    def extract_data(self,
                     ws: openpyxl.worksheet.worksheet.Worksheet,
                     sheet_name: str,
                     ) -> None:
        """
        Extract data from the worksheet and store it in the structure.
        """
        logger.debug("Extracting data from sheet: %s", sheet_name)
        column_indicies = self.get_column_indexes(ws)
        logger.debug("Column indexes: %s", column_indicies)
        for row in ws.iter_rows(min_row=2, values_only=True):
            table_name = row[column_indicies["table_name"]]
            column_name = row[column_indicies["column_name"]]
            security_classification_value = row[
                column_indicies["security_classification"]
            ]
            logger.info("table name column: %s", table_name)
            logger.info("column name column: %s", column_name)
            logger.info(
                "security classification value in row: %s",
                security_classification_value,
            )

            if not table_name or not column_name:
                logger.warning(
                    "Skipping row with missing schema, table name, or column"
                    " name: %s",
                    row,
                )
                continue

            # Convert to string for consistency
            table_name_str = str(table_name).strip()
            column_name_str = str(column_name).strip()
            logger.debug(
                "security_classification_name (column name): %s",
                security_classification_value,
            )
            self.dc.add_classification(
                table_name_str,
                column_name_str,
                security_classification_value,
            )

    def verify_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        """
        Verify that the worksheet has the required columns.

        Looks through the worksheet for the required columns:
        - TABLE_NAME
        - COLUMN_NAME
        - SECURITY_CLASSIFICATION

        Uses the get_column_indexes to identify the positions of these columns.
        If the required columns are not found, it logs an error and returns
        False.

        The get_column_indexes() method contains mapping for the various
        column names that are used in the spreadsheets encountered so far.

        :param ws: input spreadsheet worksheet to verify
        :type ws: openpyxl.worksheet.worksheet.Worksheet
        :return: boolean indicating if the worksheet has the required columns
        :rtype: bool
        """
        columns_idx = self.get_column_indexes(ws, no_raise=True)
        if len(columns_idx) != len(config.ColumnIndexKeys):
            logger.error(
                "Required columns not found in worksheet %s",
                ws.title,
            )
            return False
        for key in columns_idx:
            if columns_idx[key] is None:
                logger.error(
                    "Required column %s not found in worksheet %s",
                    key,
                    ws.title,
                )
                return False
        return True

    def get_column_indexes(
            self,
            ws: openpyxl.worksheet.worksheet.Worksheet,
            no_raise: bool=False) -> dict[str, int]:  # noqa: FBT001 FBT002
        """
        Get the column indexes for the required columns.
        """

        table_name_columns = config.TABLE_COLUMN_NAMES
        column_name_columns = config.COLUMN_COLUMN_NAMES
        security_class = config.SECURITY_CLASSIFICATION_COLUMN_NAMES

        col_idx_enum = config.ColumnIndexKeys
        column_indexes = {
            col_idx_enum.TABLE_NAME.value: None,
            col_idx_enum.COLUMN_NAME.value: None,
            col_idx_enum.SECURITY_CLASSIFICATION.value: None,
        }

        column_name_columns = [col.upper() for col in column_name_columns]
        security_class = [sc.upper() for sc in security_class]
        table_name_columns = [col.upper() for col in table_name_columns]

        for row in ws.iter_rows(max_row=1, values_only=True):
            for idx, cell_value_raw in enumerate(row):
                if cell_value_raw:
                    cell_value = cell_value_raw.upper()
                    if cell_value in table_name_columns:
                        column_indexes[col_idx_enum.TABLE_NAME.value] = idx
                    elif cell_value in column_name_columns:
                        column_indexes[col_idx_enum.COLUMN_NAME.value] = idx
                    elif cell_value in security_class:
                        column_indexes[
                            col_idx_enum.SECURITY_CLASSIFICATION.value] = idx
                        logger.debug("sec class found at index %s", idx)

        logger.debug("Column indexes: %s", column_indexes)
        if (
            column_indexes[col_idx_enum.TABLE_NAME.value] is None
            or column_indexes[col_idx_enum.COLUMN_NAME.value] is None
            or column_indexes[col_idx_enum.SECURITY_CLASSIFICATION.value]
            is None
        ):
            logger.error(
                "Required columns not found in worksheet %s",
                ws.title,
            )
            if not no_raise:
                raise ValueError
        logger.debug("Final column indexes: %s", column_indexes)
        return column_indexes

    def has_data(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        """Check if worksheet has data using openpyxl."""
        # Check if worksheet has any cells with values
        for row in ws.iter_rows():
            if row:
                if (
                    row[0].value
                    and row[1].value
                    and row[0].value.upper() == "CLASSIFICATIONS"
                    and row[1].value.upper() == "DEFINITIONS"
                ):
                    logger.debug(
                        "Skipping row with CLASSIFICATIONS and DEFINITIONS"
                        " headers",
                    )
                    # Classifications	Definitions
                    # this is an indicator that the data classfication template
                    # sheet which contains metadata about the data
                    # classification is the current sheet
                    return False
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        return True
        return False




if __name__ == "__main__":
    logging_config_file = pathlib.Path("__file__").parent / "logging.config"
    logging.config.fileConfig(logging_config_file)
    logger = logging.getLogger("main")

    src_dir = (
        pathlib.Path(__file__).parent
        / ".."
        / "project_specific"
        / "lexis"
        / "data_classifications"
    )
    dest_doc = (
        pathlib.Path(__file__).parent
        / ".."
        / "project_specific"
        / "lexis"
        / "data_classifications"
        / "consolidated_data_classification.json"
    )

    consolidator = ConsoldateDC(src_dir, dest_doc)
    consolidator.consolidate()
    # Here you would call methods to perform the consolidation
    # e.g., consolidator.consolidate()
